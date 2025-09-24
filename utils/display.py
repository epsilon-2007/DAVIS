import os
import sys
import numpy as np
import seaborn as sns
import sklearn.metrics as sk
import matplotlib.pyplot as plt

import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

recall_level_default = 0.95

def save_excel_sheet(dataset_names, fpr_list, auroc_list, hyper_parameter, file_path, template_path, sheet_name = 'Summary'):

    metrics = ['FPR', 'AUROC']
    columns = pd.MultiIndex.from_tuples(
        [(ds, m) for ds in dataset_names + ['Average'] for m in metrics] + [('Remark', '')]
    )
    
    # create a row of data as (FPR95, AUROC) per dataset
    row_data = []
    for fpr, auroc in zip(fpr_list, auroc_list):
        row_data.extend([fpr, auroc])
    row_data.append(hyper_parameter)

    df = pd.DataFrame([row_data], columns=columns)
    df.iloc[:, :-1] = df.iloc[:, :-1].applymap(lambda x: round(x * 100, 2))

    # --- Load workbook & named sheet ---
    if not os.path.exists(file_path):
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template not found at: {template_path}")
        shutil.copy(template_path, file_path)

    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}")
    ws = wb[sheet_name]
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(file_path)

def plot_distribution(args, id_scores, ood_scores, plot_distribution_file_name):
    sns.set_theme(style="white", palette="muted")
    palette = ['#A8BAE3', '#55AB83']
    # sns.displot({"ID": -1 * id_scores, "OOD":  -1 * ood_scores}, label="id", kind = "kde", palette=palette, fill = True, alpha = 0.8)
    sns.displot({"ID": id_scores, "OOD":  ood_scores}, label="id", kind = "kde", palette=palette, fill = True, alpha = 0.8)
    plt.savefig(plot_distribution_file_name, bbox_inches='tight')

def print_measures(log, auroc, aupr, fpr, method_name='simple', recall_level=recall_level_default):
    if log == None: 
        print('FPR{:d}:\t\t\t{:.2f}'.format(int(100 * recall_level), 100 * fpr))
        print('AUROC: \t\t\t{:.2f}'.format(100 * auroc))
        print('AUPR:  \t\t\t{:.2f}'.format(100 * aupr))
    else:
        log.debug('\t\t\t\t' + method_name)
        log.debug('  FPR{:d} AUROC AUPR'.format(int(100*recall_level)))
        log.debug('& {:.2f} & {:.2f} & {:.2f}'.format(100*fpr, 100*auroc, 100*aupr))

def print_measures_with_std(log, aurocs, auprs, fprs, method_name='simple', recall_level=recall_level_default):
    if log: 
        log.debug('\t\t\t\t' + method_name)
        log.debug('  FPR{:d} AUROC AUPR'.format(int(100*recall_level)))
        log.debug('& {:.2f} & {:.2f} & {:.2f}'.format(100*np.mean(fprs), 100*np.mean(aurocs), 100*np.mean(auprs)))
        log.debug('& {:.2f} & {:.2f} & {:.2f}'.format(100*np.std(fprs), 100*np.std(aurocs), 100*np.std(auprs)))
    else:
        print('FPR{:d}:\t\t\t{:.2f}\t+/- {:.2f}'.format(int(100 * recall_level), 100 * np.mean(fprs), 100 * np.std(fprs)))
        print('AUROC: \t\t\t{:.2f}\t+/- {:.2f}'.format(100 * np.mean(aurocs), 100 * np.std(aurocs)))
        print('AUPR:  \t\t\t{:.2f}\t+/- {:.2f}'.format(100 * np.mean(auprs), 100 * np.std(auprs)))


def stable_cumsum(arr, rtol=1e-05, atol=1e-08):
    """Use high precision for cumsum and check that final value matches sum
    Parameters
    ----------
    arr : array-like
        To be cumulatively summed as flat
    rtol : float
        Relative tolerance, see ``np.allclose``
    atol : float
        Absolute tolerance, see ``np.allclose``
    """
    out = np.cumsum(arr, dtype=np.float64)
    expected = np.sum(arr, dtype=np.float64)
    if not np.allclose(out[-1], expected, rtol=rtol, atol=atol):
        raise RuntimeError('cumsum was found to be unstable: '
                           'its last element does not correspond to sum')
    return out


def fpr_and_fdr_at_recall(y_true, y_score, recall_level=0.95, pos_label=None):
    classes = np.unique(y_true)
    if (pos_label is None and
            not (np.array_equal(classes, [0, 1]) or
                     np.array_equal(classes, [-1, 1]) or
                     np.array_equal(classes, [0]) or
                     np.array_equal(classes, [-1]) or
                     np.array_equal(classes, [1]))):
        raise ValueError("Data is not binary and pos_label is not specified")
    elif pos_label is None:
        pos_label = 1.

    # make y_true a boolean vector
    y_true = (y_true == pos_label)

    # sort scores and corresponding truth values
    desc_score_indices = np.argsort(y_score, kind="mergesort")[::-1]
    y_score = y_score[desc_score_indices]
    y_true = y_true[desc_score_indices]

    # y_score typically has many tied values. Here we extract
    # the indices associated with the distinct values. We also
    # concatenate a value for the end of the curve.
    distinct_value_indices = np.where(np.diff(y_score))[0]
    threshold_idxs = np.r_[distinct_value_indices, y_true.size - 1]

    # accumulate the true positives with decreasing threshold
    tps = stable_cumsum(y_true)[threshold_idxs]
    fps = 1 + threshold_idxs - tps      # add one because of zero-based indexing

    thresholds = y_score[threshold_idxs]

    recall = tps / tps[-1]

    last_ind = tps.searchsorted(tps[-1])
    sl = slice(last_ind, None, -1)      # [last_ind::-1]
    recall, fps, tps, thresholds = np.r_[recall[sl], 1], np.r_[fps[sl], 0], np.r_[tps[sl], 0], thresholds[sl]

    cutoff = np.argmin(np.abs(recall - recall_level))

    return fps[cutoff] / (np.sum(np.logical_not(y_true)))   # , fps[cutoff]/(fps[cutoff] + tps[cutoff])


def get_measures(_pos, _neg, recall_level=0.95):
    pos = np.array(_pos[:]).reshape((-1, 1))
    neg = np.array(_neg[:]).reshape((-1, 1))
    examples = np.squeeze(np.vstack((pos, neg)))
    labels = np.zeros(len(examples), dtype=np.int32)
    labels[:len(pos)] += 1

    auroc = sk.roc_auc_score(labels, examples)
    aupr = sk.average_precision_score(labels, examples)
    fpr = fpr_and_fdr_at_recall(labels, examples, recall_level)

    return fpr, aupr, auroc

def get_and_print_results(args, in_score, out_score, auroc_list, aupr_list, fpr_list, log = None):
    '''
    1) evaluate detection performance for a given out-dataset
    2) print results (FPR95, AUROC, AUPR)
    '''
    fpr, aupr, auroc = get_measures(in_score, out_score)

    # used to calculate the avg over multiple OOD test sets
    fpr_list.append(fpr)
    aupr_list.append(aupr)
    auroc_list.append(auroc) 

    print_measures(log, auroc, aupr, fpr, args.score)